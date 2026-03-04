import json
import os
import re
import textwrap
from datetime import datetime
from pathlib import Path
import tempfile

from openai import OpenAI


SUPPORTED_EXTENSIONS = {".txt", ".pdf", ".xlsx"}


def get_openai_client():
    key = os.getenv("OPENAI_API_KEY")
    if not key:
        raise RuntimeError("OPENAI_API_KEY environment variable not set")
    return OpenAI(api_key=key)


def _read_text_file(path):
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


def _read_pdf_text(path):
    try:
        from pypdf import PdfReader
    except Exception as exc:
        raise RuntimeError("pypdf is required for PDF support") from exc

    reader = PdfReader(path)
    pages = []
    for page in reader.pages:
        pages.append((page.extract_text() or "").strip())
    return "\n".join(pages).strip()


def _read_xlsx_rows(path):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("openpyxl is required for XLSX support") from exc

    wb = load_workbook(path, data_only=True)
    rows = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1):
            row_vals = []
            for cell in row:
                if cell.value is None:
                    row_vals.append("")
                else:
                    row_vals.append(str(cell.value).strip())
            rows.append((ws.title, row[0].row, row_vals))
    return rows


def parse_questionnaire(path):
    ext = Path(path).suffix.lower()
    questions = []
    structure = {"format": ext, "questions": []}

    if ext == ".pdf":
        text = _read_pdf_text(path)
        lines = text.splitlines()
        for idx, line in enumerate(lines):
            q = line.strip()
            if q:
                questions.append(q)
                structure["questions"].append({"index": len(questions) - 1, "line": idx, "text": q})
    elif ext == ".xlsx":
        rows = _read_xlsx_rows(path)
        for sheet, row_no, row_vals in rows:
            first_non_empty = None
            first_col = None
            for i, val in enumerate(row_vals, start=1):
                if val:
                    first_non_empty = val
                    first_col = i
                    break
            if first_non_empty:
                questions.append(first_non_empty)
                structure["questions"].append(
                    {
                        "index": len(questions) - 1,
                        "sheet": sheet,
                        "row": row_no,
                        "col": first_col,
                        "text": first_non_empty,
                    }
                )
    else:
        text = _read_text_file(path)
        lines = text.splitlines()
        for idx, line in enumerate(lines):
            q = line.strip()
            if q:
                questions.append(q)
                structure["questions"].append({"index": len(questions) - 1, "line": idx, "text": q})

    return "\n".join(questions), structure


def parse_reference(path):
    ext = Path(path).suffix.lower()
    if ext == ".pdf":
        return _read_pdf_text(path)
    if ext == ".xlsx":
        rows = _read_xlsx_rows(path)
        parts = []
        for sheet, row_no, row_vals in rows:
            row_text = " | ".join([v for v in row_vals if v])
            if row_text:
                parts.append(f"[{sheet} row {row_no}] {row_text}")
        return "\n".join(parts)
    return _read_text_file(path)


def generate_answers(questionnaire_text, references):
    questions = [q.strip() for q in questionnaire_text.splitlines() if q.strip()]
    ref_texts = [(ref.filename, ref.content) for ref in references]
    answers = []
    if not questions:
        return answers

    key = os.getenv("OPENAI_API_KEY")
    use_openai = bool(key)
    stopwords = {
        "the", "is", "are", "a", "an", "of", "to", "for", "in", "on", "with", "and",
        "or", "what", "which", "who", "where", "when", "why", "how", "does", "do"
    }

    for q in questions:
        q_tokens = {t for t in re.findall(r"\w+", q.lower()) if len(t) > 2 and t not in stopwords}
        scored = []
        for fname, text in ref_texts:
            t_tokens = set(re.findall(r"\w+", text.lower()))
            overlap = len(q_tokens & t_tokens)
            if q.lower() in text.lower() or overlap:
                scored.append((overlap, fname, text))

        if not scored:
            answers.append((q, "Not found in references.", "", "", 0.0))
            continue

        scored.sort(key=lambda x: x[0], reverse=True)
        _, docname, doctext = scored[0]
        prompt = (
            "Answer the question using only the reference. "
            "If unsupported, reply exactly: Not found in references.\n\n"
            f"Reference:\n{doctext}\n\nQuestion: {q}\n"
            f"Include citation in this style: ({docname})."
        )

        sentences = re.split(r"(?<=[.!?])\s+", doctext.strip())
        qwords = {w.lower().strip(".,?") for w in q.split() if len(w) > 3}
        snippet = ""
        for sent in sentences:
            words = {w.lower().strip(".,?") for w in sent.split()}
            if qwords & words:
                snippet = sent
                break
        if not snippet and sentences:
            snippet = sentences[0]

        doc_words = set(re.findall(r"\w+", doctext.lower()))
        confidence = len(qwords & doc_words) / (len(qwords) or 1)

        if not use_openai:
            answer_text = snippet or "Not found in references."
        else:
            try:
                client = get_openai_client()
                resp = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[{"role": "user", "content": prompt}],
                    max_tokens=300,
                )
                answer_text = (resp.choices[0].message.content or "").strip()
            except Exception:
                answer_text = snippet or "Not found in references."

        citation = f"({docname})" if answer_text != "Not found in references." else ""
        answers.append((q, answer_text, citation, snippet, confidence))
    return answers


def _ordered_answers(answers):
    return sorted(answers, key=lambda a: (a.question_order if a.question_order is not None else 10**9, a.id))


def _write_txt_export(outpath, source_path, answers, structure, source_blob=None):
    ordered = _ordered_answers(answers)
    raw_lines = None
    if source_blob:
        try:
            raw_lines = source_blob.decode("utf-8", errors="ignore").splitlines()
        except Exception:
            raw_lines = None
    if raw_lines is None:
        try:
            raw_lines = _read_text_file(source_path).splitlines()
        except Exception:
            raw_lines = [a.question for a in ordered]

    question_meta = structure.get("questions", [])
    answer_by_line = {}
    for i, meta in enumerate(question_meta):
        if i < len(ordered) and "line" in meta:
            answer_by_line[meta["line"]] = ordered[i]

    out_lines = []
    for idx, line in enumerate(raw_lines):
        out_lines.append(line)
        if idx in answer_by_line:
            ans = answer_by_line[idx]
            out_lines.append(f"Answer: {ans.answer or ''}")
            out_lines.append(f"Citations: {ans.citations or ''}")
            out_lines.append("")

    with open(outpath, "w", encoding="utf-8") as f:
        f.write("\n".join(out_lines).strip() + "\n")


def _write_xlsx_export(outpath, source_path, answers, structure, source_blob=None):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("openpyxl is required for XLSX export") from exc

    ordered = _ordered_answers(answers)
    temp_path = None
    if source_blob and not os.path.exists(source_path):
        fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        with open(temp_path, "wb") as f:
            f.write(source_blob)
        source_path = temp_path
    wb = load_workbook(source_path)
    for i, meta in enumerate(structure.get("questions", [])):
        if i >= len(ordered):
            break
        if not all(k in meta for k in ("sheet", "row", "col")):
            continue
        ws = wb[meta["sheet"]]
        row = int(meta["row"])
        col = int(meta["col"])
        ans = ordered[i]
        ws.cell(row=row, column=col + 1).value = ans.answer or ""
        ws.cell(row=row, column=col + 2).value = ans.citations or ""
    wb.save(outpath)
    if temp_path and os.path.exists(temp_path):
        os.remove(temp_path)


def _write_pdf_export(outpath, answers):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
    except Exception as exc:
        raise RuntimeError("reportlab is required for PDF export") from exc

    c = canvas.Canvas(outpath, pagesize=A4)
    width, height = A4
    y = height - 48
    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, y, "Questionnaire Responses")
    y -= 28

    ordered = _ordered_answers(answers)
    for ans in ordered:
        blocks = [
            ans.question,
            f"Answer: {ans.answer or ''}",
            f"Citations: {ans.citations or ''}",
            "",
        ]
        for b_idx, block in enumerate(blocks):
            lines = textwrap.wrap(block, width=110) or [""]
            for line in lines:
                if y < 40:
                    c.showPage()
                    c.setFont("Helvetica", 11)
                    y = height - 40
                if b_idx == 0:
                    c.setFont("Helvetica-Bold", 11)
                else:
                    c.setFont("Helvetica", 11)
                c.drawString(40, y, line)
                y -= 16
    c.save()


def export_document(questionnaire, answers, uploads_dir):
    ext = Path(questionnaire.filename).suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        ext = ".txt"

    source_path = os.path.join(uploads_dir, questionnaire.filename)
    source_blob = questionnaire.source_blob
    structure = {}
    if questionnaire.structure_json:
        try:
            structure = json.loads(questionnaire.structure_json)
        except json.JSONDecodeError:
            structure = {}

    export_dir = os.path.join(os.getcwd(), "exports")
    os.makedirs(export_dir, exist_ok=True)
    stem = Path(questionnaire.filename).stem
    ts = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    outpath = os.path.join(export_dir, f"export_{stem}_{ts}{ext}")

    if ext == ".xlsx":
        _write_xlsx_export(outpath, source_path, answers, structure, source_blob=source_blob)
    elif ext == ".pdf":
        _write_pdf_export(outpath, answers)
    else:
        _write_txt_export(outpath, source_path, answers, structure, source_blob=source_blob)
    return outpath
